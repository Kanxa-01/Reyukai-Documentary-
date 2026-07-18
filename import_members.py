"""
Import script for Reyukai '3rd Principle branch Total Mem' sheet into Postgres.

USAGE (run locally, not on Render):
    pip install openpyxl python-dateutil psycopg2-binary python-dotenv
    python import_members.py "Final_3rd_Branch_Shibicho_Sheets.xlsx"

Requires DATABASE_URL to be set (loaded automatically from a local .env file
if present, same as the main app).

BEFORE RUNNING:
1. Make sure app.py has been redeployed with the unique constraint removed
   from Receipt.receipt_no (real data has ~427 legitimately duplicated
   receipt numbers from years of manual entry).
2. Visit /setup/<SETUP_KEY> on your deployed app at least once so the
   `members`, `receipts`, and `users` tables already exist.
3. This only needs to be run ONCE. Running it twice will create duplicate
   members (the script does not check for existing data).

WHAT THIS DOES:
- Reads the '3rd Principle branch Total Mem' sheet (the consolidated master
  roster -- NOT the individual leader/variant sheets, which are skipped).
- Inserts every member using their original S.N as the database ID, so
  Oya No. references point at the correct row directly.
- Links each member's Oya (sponsor) after all members exist, since a few
  rows reference a *later* row or reference themselves (both are data
  errors in the source sheet -- these are skipped and reported, not guessed).
- Builds a Receipt record for every renewal cycle that has a receipt number
  or a renewal date filled in (up to ~13 cycles per member).
- Computes each member's status (Active/Expired) by comparing their most
  recent "Up to Valid" date against today.
- Prints a summary report at the end, including a list of rows with dates
  that couldn't be parsed, so you can fix those by hand afterward.
"""
import os
import re
import sys
from datetime import datetime, date, timedelta
from collections import Counter

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

SHEET_NAME = '3rd Principle branch Total Mem'
BATCH_SIZE = 500
MAX_RETRIES = 5
TODAY = date.today()

# Column indices (0-based) in the master sheet
COL_SN = 1
COL_FULL_NAME = 2
COL_ADDRESS = 3
COL_OYA_NAME = 4       # not used directly -- oya_name is derived from oya_id
COL_OYA_NO = 5
COL_ENTRANCE_DATE = 6
COL_HOZASHU = 7
COL_JUN_SHIBICHO = 8
COL_SHIBICHO = 9

# (receipt_no_col, renewal_date_col, valid_upto_col_or_None) for each renewal cycle
CYCLES = [
    (10, 11, None),
    (12, 13, 14),
    (15, 16, 17),
    (18, 19, 20),
    (21, 22, 23),
    (24, 25, 26),
    (27, 28, 29),
    (30, 31, 32),
    (33, 34, 35),
    (36, 37, 38),
    (39, 40, 41),
    (42, 43, 44),
    (45, 46, 47),
]


def parse_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        s = re.sub(r'\([^)]*\)', '', s).strip()
        s = re.sub(r'\s*\d+\s*(yrs?|years?)\.?\s*', '', s, flags=re.IGNORECASE).strip()
        s = re.sub(r'\bDce\b', 'Dec', s, flags=re.IGNORECASE)
        if not s:
            return None
        fmts = ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d %b %Y', '%Y %b %d', '%b %d, %Y',
                '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
        for fmt in fmts:
            try:
                d = datetime.strptime(s, fmt).date()
                if 1960 <= d.year <= 2035:
                    return d
            except ValueError:
                continue
        try:
            from dateutil import parser as dateutil_parser
            d = dateutil_parser.parse(s, fuzzy=True).date()
            if 1960 <= d.year <= 2035:
                return d
        except Exception:
            pass
    return None


def receipt_no_str(v):
    if v is None or v == '':
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def main():
    if len(sys.argv) < 2:
        print('Usage: python import_members.py <path_to_excel_file.xlsx>')
        sys.exit(1)
    excel_path = sys.argv[1]

    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        print('ERROR: DATABASE_URL is not set. Add it to a local .env file or export it first.')
        sys.exit(1)
    database_url = database_url.replace('postgres://', 'postgresql://', 1)

    print(f'Loading {excel_path} ...')
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f'ERROR: sheet "{SHEET_NAME}" not found. Available sheets: {wb.sheetnames}')
        sys.exit(1)
    ws = wb[SHEET_NAME]

    print('Connecting to database...')
    connect_kwargs = dict(
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=5,
        connect_timeout=15,
    )
    conn = psycopg2.connect(database_url, **connect_kwargs)
    conn.autocommit = False
    cur = conn.cursor()

    def reconnect():
        nonlocal conn, cur
        try:
            cur.close()
            conn.close()
        except Exception:
            pass
        conn = psycopg2.connect(database_url, **connect_kwargs)
        conn.autocommit = False
        cur = conn.cursor()

    def run_with_retry(fn):
        """Retries a DB operation on connection drops, reconnecting between tries."""
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                return fn()
            except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
                print(f'  [connection issue, attempt {attempt}/{MAX_RETRIES}: {e}]')
                if attempt == MAX_RETRIES:
                    raise
                import time
                time.sleep(3 * attempt)
                reconnect()

    print('Checking for already-imported members (so this script can resume safely if interrupted)...')
    cur.execute('SELECT id FROM members')
    existing_ids = set(r[0] for r in cur.fetchall())
    if existing_ids:
        print(f'  Found {len(existing_ids)} members already imported -- these rows will be skipped.')

    members_batch = []
    receipts_batch = []
    oya_pairs = []  # (member_id, oya_id)
    inserted_ids = set(existing_ids)

    stats = Counter()
    unparsed_dates = []  # (sn, column_index, raw_value)

    def _do_flush_members():
        execute_values(cur, """
            INSERT INTO members
                (id, full_name, address, entrance_date, hozashu, jun_shibicho,
                 shibicho, valid_upto, status, notes, created_at, updated_at)
            VALUES %s
            ON CONFLICT (id) DO NOTHING
        """, members_batch)
        conn.commit()

    def flush_members():
        if not members_batch:
            return
        run_with_retry(_do_flush_members)
        members_batch.clear()

    def _do_flush_receipts():
        execute_values(cur, """
            INSERT INTO receipts
                (receipt_no, member_id, full_name_snapshot, shibicho_snapshot,
                 entrance_date_snapshot, renewal_date, status, created_at)
            VALUES %s
        """, receipts_batch)
        conn.commit()

    def flush_receipts():
        if not receipts_batch:
            return
        flush_members()  # guarantee referenced members are committed first
        run_with_retry(_do_flush_receipts)
        receipts_batch.clear()

    print('Reading rows...')
    row_num = 0
    for row in ws.iter_rows(min_row=2, max_row=34000, max_col=49, values_only=True):
        sn = row[COL_SN]
        if sn is None:
            continue
        try:
            sn = int(sn)
        except (ValueError, TypeError):
            continue
        row_num += 1

        if sn in existing_ids:
            stats['skipped_already_imported'] += 1
            continue

        full_name = (row[COL_FULL_NAME] or '').strip() if row[COL_FULL_NAME] else ''
        if not full_name:
            stats['skipped_blank_name'] += 1
            continue

        address = (row[COL_ADDRESS] or '').strip() if row[COL_ADDRESS] else None

        entrance_date = parse_date(row[COL_ENTRANCE_DATE])
        if row[COL_ENTRANCE_DATE] not in (None, '') and entrance_date is None:
            unparsed_dates.append((sn, 'entrance_date', repr(row[COL_ENTRANCE_DATE])))

        hozashu = str(row[COL_HOZASHU]).strip() if row[COL_HOZASHU] not in (None, '') else None
        jun_shibicho = str(row[COL_JUN_SHIBICHO]).strip() if row[COL_JUN_SHIBICHO] not in (None, '') else None

        shibicho = (row[COL_SHIBICHO] or '').strip() if row[COL_SHIBICHO] else ''
        if not shibicho:
            shibicho = 'Unknown'
            stats['defaulted_shibicho'] += 1

        # Find the most recent "Up to Valid" date across all cycles
        last_valid_upto = None
        for receipt_idx, renew_idx, valid_idx in CYCLES:
            if valid_idx is not None:
                vu = parse_date(row[valid_idx])
                if row[valid_idx] not in (None, '') and vu is None:
                    unparsed_dates.append((sn, f'valid_upto_col{valid_idx}', repr(row[valid_idx])))
                if vu and (last_valid_upto is None or vu > last_valid_upto):
                    last_valid_upto = vu

        status = 'Active' if (last_valid_upto is None or last_valid_upto >= TODAY) else 'Expired'

        members_batch.append((
            sn, full_name, address, entrance_date, hozashu, jun_shibicho,
            shibicho, last_valid_upto, status, None, datetime.utcnow(), datetime.utcnow()
        ))
        inserted_ids.add(sn)
        stats['members_inserted'] += 1

        # Oya (sponsor) link -- resolved in a second pass after all members exist
        oya_no_raw = row[COL_OYA_NO]
        if oya_no_raw not in (None, '', 0, '0'):
            try:
                oya_int = int(oya_no_raw)
                if oya_int == sn:
                    stats['oya_self_ref_skipped'] += 1
                elif oya_int < 1 or oya_int > 33672:
                    stats['oya_out_of_range_skipped'] += 1
                else:
                    oya_pairs.append((oya_int, sn))
            except (ValueError, TypeError):
                stats['oya_unparseable_skipped'] += 1

        # Receipts -- one per renewal cycle that has a receipt no. or a date
        for receipt_idx, renew_idx, valid_idx in CYCLES:
            rno = receipt_no_str(row[receipt_idx])
            rdate = parse_date(row[renew_idx])
            if row[renew_idx] not in (None, '') and rdate is None:
                unparsed_dates.append((sn, f'renewal_date_col{renew_idx}', repr(row[renew_idx])))
            if rno is None and rdate is None:
                continue
            if rno is None:
                rno = f'UNKNOWN-{sn}-{receipt_idx}'
                stats['receipts_missing_no'] += 1
            receipts_batch.append((
                rno, sn, full_name, shibicho, entrance_date, rdate, status, datetime.utcnow()
            ))
            stats['receipts_inserted'] += 1

        if len(members_batch) >= BATCH_SIZE:
            flush_members()
        if len(receipts_batch) >= BATCH_SIZE:
            flush_receipts()

        if row_num % 5000 == 0:
            print(f'  ... {row_num} rows processed')

    flush_members()
    flush_receipts()

    print(f'Linking Oya (sponsor) relationships...')
    valid_oya_pairs = [(oya_id, member_id) for oya_id, member_id in oya_pairs if oya_id in inserted_ids]
    stats['oya_target_missing_skipped'] = len(oya_pairs) - len(valid_oya_pairs)
    print(f'  {len(valid_oya_pairs)} valid links ({stats["oya_target_missing_skipped"]} skipped -- sponsor row was blank-name and not imported)')
    if valid_oya_pairs:
        execute_values(cur, """
            UPDATE members AS m SET oya_id = v.oya_id
            FROM (VALUES %s) AS v(oya_id, id)
            WHERE m.id = v.id
        """, valid_oya_pairs)
        conn.commit()

    print('Resetting the members ID sequence so future app inserts continue correctly...')
    cur.execute("SELECT setval(pg_get_serial_sequence('members', 'id'), (SELECT MAX(id) FROM members))")
    conn.commit()

    cur.close()
    conn.close()

    print()
    print('=' * 60)
    print('IMPORT COMPLETE')
    print('=' * 60)
    print(f"Members inserted:              {stats['members_inserted']}")
    print(f"Skipped -- already imported:   {stats['skipped_already_imported']}")
    print(f"Receipts inserted:             {stats['receipts_inserted']}")
    print(f"  (of which missing receipt#): {stats['receipts_missing_no']}")
    print(f"Skipped -- blank name:         {stats['skipped_blank_name']}")
    print(f"Shibicho defaulted 'Unknown':  {stats['defaulted_shibicho']}")
    print(f"Oya skipped -- self-reference: {stats['oya_self_ref_skipped']}")
    print(f"Oya skipped -- out of range:   {stats['oya_out_of_range_skipped']}")
    print(f"Oya skipped -- unparseable:    {stats['oya_unparseable_skipped']}")
    print(f"Oya skipped -- sponsor blank-name: {stats['oya_target_missing_skipped']}")
    print(f"Unparseable dates:             {len(unparsed_dates)}")

    if unparsed_dates:
        report_path = 'unparsed_dates_report.txt'
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('S.N\tField\tRaw Value\n')
            for sn, field, raw in unparsed_dates:
                f.write(f'{sn}\t{field}\t{raw}\n')
        print(f'\nA list of the {len(unparsed_dates)} unparseable dates was saved to: {report_path}')
        print('These members were still imported -- just with that specific date left blank.')
        print('Fix these manually in the app afterward if needed.')


if __name__ == '__main__':
    main()
